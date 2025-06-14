import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# loading the excel sheet
book = openpyxl.load_workbook('student_record.xlsx')

#choosing the sheet from excel 
print(book)
sheet = book['Sheet1']
print(sheet)

# counting number of rows 
r = sheet.max_row
print(r)

#for looping keeping the variable 
resp = 1

# counting number of columns / subjects
c = sheet.max_column
print(c)


# list of students to remind
l1 = []
print(l1)


# to concatenate list of roll numbers with
# lack of attendance
l2 = ""
print(l2)

# list of roll numbers with lack of attendance
l3 = []
print(l3)


# staff mail ids
staff_mails = ['jay19chavan@gmail.com', 'swarajsolanke2702@gmail.com']

# Warning messages
m1 = "warning!!! you can take only one more day leave for Data Engineering  class"
m2 = "warning!!! you can take only one more day leave for Machine learning class"
m3 = "warning!!! you can take only one more day leave for python class"
m4= "warning!!! you can take only one more day leave for Computer vision class"

print(staff_mails)

# saving the data into excel sheet 
def savefile():
    book.save(r'student_record.xlsx')
    print("saved!")
savefile()
print(book)



def check(no_of_days, row_num, b):

    # to use the globally declared lists and strings
    global staff_mails
    global l2
    global l3

    for student in range(0, len(row_num)):
        # if total no.of.leaves equals threshold
        if no_of_days[student] == 2:
            if b == 1:
                
                # mail_id appending
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m1)  # sending mail
            elif b == 2:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m2)
            elif b == 3:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m3)

            else:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m4)

        # if total.no.of.leaves > threshold
        elif no_of_days[student] > 2:
            if b == 1:

                # adding roll no
                l2 = l2+str(sheet.cell(row=row_num[student], column=1).value)

                # student mail_id appending
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "Data Engineering"  # subject based on the code number

            elif b == 2:
                l2 = l2+str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "Machine learning"
            elif b == 3:
                l2 = l2+str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "python"

            else:
                l2 = l2+str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "Computer vision"

        # If threshold crossed, modify the message
        if l2 != "" and len(l3) != 0:

            # message for student
            msg1 = "you have lack of attendance in " + subject + " !!!"

            # message for staff
            msg2 = "the following students have lack of attendance in your subject : "+l2

            mailstu(l3, msg1)  # mail to students
            staff_id = staff_mails[b-1]  # pick respective staff's mail_id
            mailstaff(staff_id, msg2)  # mail to staff



# for students
def mailstu(li, msg):
    from_id = 'amrutawedding7@gmail.com'
    print(from_id)
    pwd = 'dsll hybw yqct hwia'
    print(pwd)
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    print(s)
    s.starttls()
    s.login(from_id, pwd)
    print(s)

    # for each student to warn send mail
    for i in range(0, len(li)):
        to_id = li[i]
        message = MIMEMultipart()
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg, 'plain'))
        content = message.as_string()
        s.sendmail(from_id, to_id, content)
        s.quit()
    print("mail sent to students")


# for staff
def mailstaff(mail_id, msg):
    from_id = 'jay19chavan@gmail.com'
    pwd = 'qeqw txlt xmcx kkah'
    to_id = mail_id
    message = MIMEMultipart()
    message['Subject'] = 'Lack of attendance report'
    message.attach(MIMEText(msg, 'plain'))
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
    content = message.as_string()
    s.sendmail(from_id, to_id, content)
    s.quit()
    print('Mail Sent to staff')

while resp is 1:
    print("1--->DE\n2--->ML\n3--->python\n4--->CV")

    # enter the correspondingnumber
    y = int(input("enter subject :"))

    # no.of.absentees for that subject
    no_of_absentees = int(input('no.of.absentees :'))

    if(no_of_absentees > 1):
        x = list(map(int, (input('roll nos :').split(' '))))
    else:
        x = [int(input('roll no :'))]

    # list to hold row of the student in Excel sheet
    row_num = []

    # list to hold total no.of leaves
    # taken by ith student
    no_of_days = []

    for student in x:

        for i in range(2, r+1):

            if y == 1:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=3).value
                    m = m+1
                    sheet.cell(row=i, column=3).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)

            elif y == 2:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=4).value
                    m = m+1
                    sheet.cell(row=i, column=4).value = m+1
                    no_of_days.append(m)
                    row_num.append(i)

            elif y == 3:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=5).value
                    m = m+1
                    sheet.cell(row=i, column=5).value = m+1
                    row_num.append(i)
                    no_of_days.append(m)
            elif y == 4:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=5).value
                    m = m+1
                    sheet.cell(row=i, column=5).value = m+1
                    row_num.append(i)
                    no_of_days.append(m)

    check(no_of_days, row_num, y)
    resp = int(input('another subject ? 1---->yes 0--->no'))